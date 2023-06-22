from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.views import LoginView, PasswordChangeView
from django.contrib.auth.models import User
from django.views.generic.edit import UpdateView, CreateView, DeleteView
from django.views.generic import DetailView, View
from django.utils.encoding import smart_str
from django.core.mail import send_mail
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from .models import *
from .forms import *

# Create your views here.



def error404(request, exception) :
    
    return render(request, 'intern/App/404.html', status=404)    
    


class CustomLoginView(LoginView) :
    
    template_name = 'intern/Login.html'
    redirect_authenticated_user = True
    success_url = reverse_lazy('home')
    
    def form_invalid(self, form) : 
        
        context = self.get_context_data()
        context['form'] = form
        context['error'] = 'Nom d\'utilisateur et/ou mot de passe incorrect(s).'
        return self.render_to_response(context)



@login_required
def home(request) :
    
    user = request.user
    intern = Intern.objects.get(user=user)
    internship = Intership.objects.get(intern=intern, status='En cours')
    project = Project.objects.get(internship=internship, status='En cours')
    Tasks = Task.objects.all()
    internships = intern.intership_set.all()
    
    Projects = Project.objects.all()
    
    recent_tasks = project.task_set.order_by('-id')[:5] # Liste des tâches récentes.
    recent_documents =  project.document_set.order_by('-id')[:5] # Liste des documents récents.
    
    data = [] # Les données pour le graphe 'Tâches'.
    data_1 = [] # Les données pour le graphe 'Projets'.
    
    Annule = []
    EnAttente = []
    EnCours = []
    Termine = []
    
    P_Annule = []
    P_EnAttente = []
    P_EnCours = []
    P_Termine = []
    
    for task in Tasks : 
        
        if task.status == 'Annulé' and task.project.internship == internship : 
            
            Annule.append(task)
            
        elif task.status == 'En attente' and task.project.internship == internship : 
            
            EnAttente.append(task)
            
        elif task.status == 'En cours' and task.project.internship == internship : 
            
            EnCours.append(task)
            
        elif task.status == 'Terminé' and task.project.internship == internship : 
            
            Termine.append(task)
        
    data.append(len(Annule))
    data.append(len(EnAttente))
    data.append(len(EnCours))
    data.append(len(Termine))
    
    for internship in internships : 
        
        project = internship.project_set.all()
        
        for project in project : 
        
            if project.status == 'Annulé' : 
                
                P_Annule.append(project)
                
            elif project.status == 'En attente' : 
                
                P_EnAttente.append(project)
                
            elif project.status == 'En cours' : 
                
                P_EnCours.append(project)
                
            elif project.status == 'Terminé' : 
                
                P_Termine.append(project)
    
    data_1.append(len(P_Annule))
    data_1.append(len(P_EnAttente))
    data_1.append(len(P_EnCours))
    data_1.append(len(P_Termine))
    
    return render(request, 'intern/App/index.html', {'Projects' : Projects, 'recent_tasks' : recent_tasks, 'recent_documents' : recent_documents, 'data' : data, 'data_1' : data_1})



@login_required
def project(request) :
    
    Projects = Project.objects.all()
    Interships = Intership.objects.all()
    Documents = Document.objects.all()

    return render(request, 'intern/App/pages/project.html', {'Projects' : Projects, 'Interships' : Interships, 'Documents' : Documents})



@login_required
def tasklist(request) :
    
    Projects = Project.objects.all()
    Tasks = Task.objects.order_by('status', 'start_date')

    return render(request, 'intern/App/pages/taskboard.html', {'Projects' : Projects,'Tasks' : Tasks})



@login_required
def tasklistEnded(request, id) :
    
    project = Project.objects.get(id=id)
    Tasks = Task.objects.filter(project=project).order_by('status', 'start_date')

    return render(request, 'intern/App/pages/taskboardEnded.html', {'project' : project,'Tasks' : Tasks})



@login_required
def profile_detail_and_update(request) :
    
    Interships = Intership.objects.all()
    user = request.user
    form = UpdateProfileForm(request.POST or None, instance=user)
    
    if request.method == 'POST':
        
        if form.is_valid() :
            
            form.save()
            return redirect('profile')
    
    return render(request, 'intern/App/pages/profile.html', {'Interships' : Interships, 'form' : form})



@login_required
def documentlist(request) :
    
    Documents = Document.objects.order_by('title')

    return render(request, 'intern/App/pages/documents.html', {'Documents' : Documents})



@login_required
def uploadDocument(request) : 
    
    form = DocumentUploadForm(request.POST, request.FILES)
    
    if request.method == 'POST' :
        
        # form = DocumentUploadForm(request.POST, request.FILES)
        
        if form.is_valid():
           
            form.save()
            return redirect('documents')
    
    return render(request, 'intern/App/pages/document_add.html', {'form' : form}) 



class UploadDocumentView(LoginRequiredMixin, CreateView) :

    template_name = 'intern/App/pages/document_add.html'
    model = Task
    form_class = DocumentUploadForm
    success_url = reverse_lazy('documents')
    
    # On definit la methode dispatch, pour pouvoir utiliser l'objet request.
    def dispatch(self, request, *args, **kwargs):
        self.request = request
        return super().dispatch(request, *args, **kwargs)
    
    # Pour définir le projet de l'utilisateur dynamiquement.
    # (Ici, on recupere l'utilisateur et son projet 'En cours').
    def get_initial(self) :
        
        initial = super(UploadDocumentView, self).get_initial()
        user = self.request.user # On recupere l'utilisateur actuellement connecté.
        intern = Intern.objects.get(user=user)
        internships = Intership.objects.filter(intern=intern)
        
        for internship in internships :
            
            if internship.status == 'En cours' :
                
                project = Project.objects.get(internship=internship)
                initial['project'] = project
        
        return initial   
    
    # Pour rendre le champ de choix du projet invisible.
    def get_form(self, form_class=None) :
        
        form = super().get_form(form_class)
        form.fields['project'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['phase'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['task'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['project'].label = ''
        form.fields['phase'].label = ''
        form.fields['task'].label = ''
        return form
        
        
        
class UploadDocumentToTaskView(LoginRequiredMixin, CreateView) :

    template_name = 'intern/App/pages/document_add_to_task.html'
    model = Document
    form_class = DocumentUploadForm
    success_url = reverse_lazy('documents')
    
    # On definit la methode dispatch, pour pouvoir utiliser l'objet request.
    def dispatch(self, request, *args, **kwargs):
        self.request = request
        return super().dispatch(request, *args, **kwargs)
    
    # Pour définir le projet de l'utilisateur dynamiquement.
    # (Ici, on recupere l'utilisateur et son projet 'En cours').
    def get_initial(self) :
        
        initial = super(UploadDocumentToTaskView, self).get_initial()
        user = self.request.user # On recupere l'utilisateur actuellement connecté.
        intern = Intern.objects.get(user=user)
        internships = Intership.objects.filter(intern=intern)
        
        for internship in internships :
            
            if internship.status == 'En cours' :
                
                project = Project.objects.get(internship=internship)
                initial['project'] = project
        
        return initial   
    
    # Pour rendre le champ de choix du projet invisible.
    def get_form(self, form_class=None) :
        
        form = super().get_form(form_class)
        form.fields['project'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['phase'].widget = forms.HiddenInput(attrs={'hidden':True})
        # form.fields['task'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['project'].label = ''
        form.fields['phase'].label = ''
        form.fields['task'].label = 'Tâche'
        return form
    
    
    
class UpdateDocumentView(LoginRequiredMixin, UpdateView) :

    template_name = 'intern/App/pages/document_update.html'
    model = Document
    form_class = UpdateDocumentForm
    success_url = reverse_lazy('documents')
    
    def get_form(self, form_class=None) :
        
        form = super().get_form(form_class)
        form.fields['task'].label = 'Tâche'
        return form



class ProfilePhotoUpdate(LoginRequiredMixin, UpdateView) :
    
    model = Intern
    template_name = 'intern/App/pages/profile_photo_update.html'
    form_class = UpdatePhotoForm
    success_url = reverse_lazy('profile')



# @login_required
# def profile_photo_update(request) :
    
#     form = UpdatePhotoForm(request.POST)
    
#     if request.method == 'POST':
        
#         if form.is_valid() :
            
#             form = UpdatePhotoForm(request.POST, request.FILES)
#             photo = form.save(commit=False)
#             # set the uploader to the user before saving the model
#             photo.uploader = request.user
#             # now we can save
#             photo.save()
#             return redirect('profile')
    
#     return render(request, 'intern/App/pages/profile_photo_update.html', {'form' : form})



class ChangePasswordView(LoginRequiredMixin, PasswordChangeView) :
    
    template_name = 'intern/App/pages/password_change.html'
    success_url = reverse_lazy('profile')
    
    def form_valid(self, form):
        
        messages.success(self.request, 'Votre mot de passe a été modifié avec succès.')
        return super().form_valid(form)



class DetailledTaskView(LoginRequiredMixin, UpdateView) :

    template_name = 'intern/App/pages/task_detailled.html'
    model = Task
    form_class = UpdateTaskForm
    success_url = reverse_lazy('tasklist')
    
    def get_context_data(self, **kwargs) :
        
        context = super().get_context_data(**kwargs)
        Documents = Document.objects.all()
        my_dict = {'Documents': Documents}
        context.update(my_dict)
        return context
    
    
    
class AddTaskView(LoginRequiredMixin, CreateView) :

    template_name = 'intern/App/pages/task_add.html'
    model = Task
    form_class = AddTaskForm
    success_url = reverse_lazy('tasklist')
    
    # On definit la methode dispatch, pour pouvoir utiliser l'objet request.
    def dispatch(self, request, *args, **kwargs):
        self.request = request
        return super().dispatch(request, *args, **kwargs)
    
    # Pour définir le projet de l'utilisateur dynamiquement.
    # (Ici, on recupere l'utilisateur et son projet 'En cours').
    def get_initial(self) :
        
        initial = super(AddTaskView, self).get_initial()
        user = self.request.user # On recupere l'utilisateur actuellement connecté.
        intern = Intern.objects.get(user=user)
        # internship = Intership.objects.get(intern=intern)
        internships = Intership.objects.filter(intern=intern)
        # project = Project.objects.get(internship=internship)
        
        for internship in internships :
            
            if internship.status == 'En cours' :
                
                project = Project.objects.get(internship=internship)
                initial['project'] = project
        
        initial['status'] = 'En attente'
        
        return initial
    
    # Pour rendre le champ de choix du projet invisible.
    def get_form(self, form_class=None) :
        
        form = super().get_form(form_class)
        form.fields['project'].widget = forms.HiddenInput(attrs={'hidden':True})
        form.fields['project'].label = ''
        return form
    
    
    
class DeleteTask(LoginRequiredMixin, DeleteView) :
    
    template_name = 'intern/App/pages/task_delete.html'
    model = Task
    success_url = reverse_lazy('tasklist')
    
    
    
class UpdateProjectView(LoginRequiredMixin, UpdateView) : 
    
    model = Project
    template_name = 'intern/App/pages/project_update.html'
    form_class = UpdateProjectForm
    success_url = reverse_lazy('project')
    
    
    
class DetailledProjectView(LoginRequiredMixin, DetailView) :
    
    model = Project
    template_name = 'intern/App/pages/project_detailled.html'
    
    def get_context_data(self, **kwargs) :
        
        context = super().get_context_data(**kwargs)
        Documents = Document.objects.all()
        my_dict = {'Documents': Documents}
        context.update(my_dict)
        return context
    
    
    
class InfoTaskView(LoginRequiredMixin, DetailView) :
    
    model = Task
    template_name = 'intern/App/pages/task_info.html'
    
    def get_context_data(self, **kwargs) :
        
        context = super().get_context_data(**kwargs)
        Documents = Document.objects.all()
        my_dict = {'Documents': Documents}
        context.update(my_dict)
        return context
    
    
    
class InfoDocumentView(LoginRequiredMixin, DetailView) :
    
    model = Document
    template_name = 'intern/App/pages/document_info.html'
    
    
    
class DownloadFileView(View) :
    
    def get(self, request, pk) :
        
        document = get_object_or_404(Document, pk=pk)
        file_path = document.fichier.path
        
        with open(file_path, 'rb') as file :
            
            response = HttpResponse(file.read(), content_type='application/force-download')
            response['Content-Disposition'] = 'attachment; filename=%s' % smart_str(document.fichier.name.split('/')[-1])
            
            return response
        
        
        
class DeleteDocument(LoginRequiredMixin, DeleteView) :
    
    template_name = 'intern/App/pages/document_delete.html'
    model = Document
    success_url = reverse_lazy('documents')
    
    
    
class SearchView(View) :
    
    def get(self, request) :
        
        query = request.GET.get('q')
        results = []
        results_projects = []
        results_tasks = []
        results_docs = []
        
        if query :
            
            # Effectuer la recherche dans les différents modèles
            model1_results = Project.objects.filter(Q(title__icontains=query) | Q(description__icontains=query))
            model2_results = Task.objects.filter(Q(title__icontains=query) | Q(description__icontains=query))
            model3_results = Document.objects.filter(Q(title__icontains=query) | Q(description__icontains=query))
            # Ajouter les résultats de chaque modèle à la liste des résultats
            results.extend(list(model1_results))
            results.extend(list(model2_results))
            results.extend(list(model3_results))
            
            # On trie les résultats de la recherche en foncftions des instances.
            for result in results : 
                
                if isinstance(result, Project) :
                    
                    results_projects.append(result)
                    
                elif isinstance(result, Task) :
                    
                    results_tasks.append(result)
                    
                if isinstance(result, Document) :
                    
                    results_docs.append(result)
            
        return render(request, 'intern/App/pages/search.html', {'results_projects' : results_projects, 'results_tasks' : results_tasks, 'results_docs' : results_docs, 'results': results, 'query': query})
    
    
    
def send_welcome_email(request):
    

    user_id = request.GET.get('user_id', '')
    user = User.objects.get(pk=user_id)

    if user:
        
        subject = 'Bienvenue chez ICCSOFT pour votre stage.'
        message = 'Vous avez été inscrit sur IPM(Intership Project Manager), l\'application web utilisé pour gérer'
        message += '\n les projets de stage chez ICCSOFT. Vos identfiants de connexion sont : '
        message += f'\n Nom d\'utisateur : {user.username} \n Mot de passe : S3cret1234! '
        message += '\n l\'application est disponible sur : localhost:8000/ '
        message += '\nVous pouvez changez vos informations personnelles et votre mot de passe sur la page correspondant à votre profil.'
        message += '\n Nous vous souhaitons une bonne période de stage.'
        from_email = 'mbohlulajonathan4@gmail.com' 
        recipient_list = [user.email]
        send_mail(subject, message, from_email, recipient_list)
        messages.success(request, 'Un email avec un lien de connexion a été envoyé à {}.'.format(user.email))

        try:
            obj = User.objects.get(pk=user.pk)
            return redirect(obj)
        except:
            return redirect('/admin')
        
        
        
def export_to_excel(request, id) :
    
    user = request.user
    intern = Intern.objects.get(user=user)
    internship = Intership.objects.get(intern=intern, id=id)
    project = Project.objects.get(internship=internship)
    Tasks = Task.objects.all()
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachement; filename=Tâches - Projet({project.title}).xlsx'
    
    # On crée un nouveau classeur Excel.
    wb = Workbook()

    # On crée une nouvelle feuille de calcul. (celle des tâches)
    ws = wb.active
    ws.title = "Tâches"
    
    # On crée une nouvelle feuille de calcul. (celle du projet)
    ws_1 = wb.create_sheet(title="Projet")
    # ws_1.title = "Projet"

    # On récupère les données du modèle.
    data_1 = []
    
    for task in Tasks :
        
        if task.project.title == project.title :
            
            data_1.append(task)

    # On écrit les données dans la feuille de calcul.
    columns = ['Titre', 'Description', 'Date de début', 'Date de fin', 'État']
    row_num = 1
    
    # First sheet (celle des tâches).
    for col_num, column_title in enumerate(columns, 1) :
        
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    for obj in data_1 :
        
        row_num += 1
        row = [obj.title, obj.description, obj.start_date, obj.end_date, obj.status]
        for col_num, cell_value in enumerate(row, 1) :
        
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # Second sheet (celle du projet).
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1) :
        
        cell = ws_1.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    row_num += 1       
    row_1 = [project.title, project.description, project.start_date, project.end_date, project.status]
    for col_num, cell_value in enumerate(row_1, 1) :

        cell = ws_1.cell(row=row_num, column=col_num)
        cell.value = cell_value

    wb.save(response)
    return response



def export_to_excel_current_project(request) :
    
    user = request.user
    intern = Intern.objects.get(user=user)
    internship = Intership.objects.get(intern=intern, status='En cours')
    project = Project.objects.get(internship=internship, status='En cours')
    Tasks = Task.objects.all()
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachement; filename=Tâches - Projet({project.title}).xlsx'
    
    # On crée un nouveau classeur Excel.
    wb = Workbook()

    # On crée une nouvelle feuille de calcul. (celle des tâches)
    ws = wb.active
    ws.title = "Tâches"
    
    # On crée une nouvelle feuille de calcul. (celle du projet)
    ws_1 = wb.create_sheet(title="Projet")
    # ws_1.title = "Projet"

    # On récupère les données du modèle.
    data_1 = []
    
    for task in Tasks :
        
        if task.project.title == project.title :
            
            data_1.append(task)

    # On écrit les données dans la feuille de calcul.
    columns = ['Titre', 'Description', 'Date de début', 'Date de fin', 'État']
    row_num = 1
    
    # First sheet (celle des tâches).
    for col_num, column_title in enumerate(columns, 1) :
        
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    for obj in data_1 :
        
        row_num += 1
        row = [obj.title, obj.description, obj.start_date, obj.end_date, obj.status]
        for col_num, cell_value in enumerate(row, 1) :
        
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = cell_value
    
    # Second sheet (celle du projet).
    row_num = 1
    
    for col_num, column_title in enumerate(columns, 1) :
        
        cell = ws_1.cell(row=row_num, column=col_num)
        cell.value = column_title
    
    row_num += 1       
    row_1 = [project.title, project.description, project.start_date, project.end_date, project.status]
    for col_num, cell_value in enumerate(row_1, 1) :

        cell = ws_1.cell(row=row_num, column=col_num)
        cell.value = cell_value

    wb.save(response)
    return response



def search(request) : 
    
    query = request.GET.get('q')
    user = request.user
    
    # Pour éviter d'effectuer une requête avec une valeur 'None'.
    if query is not None : 
        
        results = Project.objects.filter(Q(title__icontains=query) | Q(description__icontains=query))
        context = {
            'results' : results,
            'query' : query,
        }
    
        return render(request, 'intern/App/pages/search.html', context)
    
    return render(request, 'intern/App/pages/search.html')
